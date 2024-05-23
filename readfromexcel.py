import openpyxl

# Use raw string or double backslashes for the path
path = "C:\\Users\Spanidea-LT-103\Desktop\YASHSPANIDEA\Book1.xlsx"

# Load the workbook
workbook = openpyxl.load_workbook(path)

# Access the specific sheet by name
sheet = workbook["Sheet1"]

# Get the maximum number of rows and columns
rows = sheet.max_row
cols = sheet.max_column

# Print the number of rows and columns
print(f"Rows: {rows}")
print(f"Columns: {cols}")

# Loop through each cell in the sheet and print its value
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        value = sheet.cell(row=r, column=c).value
        print(value, end="   ")
    print()  # Move to the next line after each row
